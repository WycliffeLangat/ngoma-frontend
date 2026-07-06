from collections import defaultdict
from datetime import datetime
from pathlib import Path
import re
import unicodedata

from django.utils.text import slugify
from django.db.models import Sum
from openpyxl import load_workbook

from charts.artist_metadata import artist_country
from charts.methodology import (
    CERTIFICATION_THRESHOLDS,
    public_points,
)


# ---------------------------------------------------------------------------
# Merge-workbook utilities
# Shared by import_master_workbook() (DB seed) and export_frontend_data.py
# (static JS export) so the same canonicalization runs at both entry points.
# ---------------------------------------------------------------------------

def _broad_normalize(text):
    """
    Broad ASCII normalization used to fuzzy-match merge-workbook entries
    against raw master data.  Strips accents, lowercases, removes punctuation
    and common stop words so minor spelling variants resolve to the same key.
    """
    if not text:
        return ""
    s = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode()
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\b(ft|feat|featuring|and|the|a|an|of|in|on|by)\b", " ", s)
    return " ".join(s.split())


def load_merge_pairs(merge_workbook_path):
    """
    Read approved merge pairs from the duplicate-release merge workbook.

    Returns a dict:
        {(broad_title, broad_artist, chart_type): (canonical_title, canonical_artist)}

    Both the duplicate and the keeper share the same Title/Artist(s)/Type in the
    "Apply Merges" sheet — those are the canonical display strings to use everywhere.
    """
    wb = load_workbook(merge_workbook_path, read_only=True, data_only=True)
    if "Apply Merges" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Apply Merges"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    pairs = {}
    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        r = dict(zip(headers, row))
        # Honour an explicit Action column if present; otherwise treat every row
        # in "Apply Merges" as a merge (that is the purpose of the sheet).
        action = str(r.get("Action") or "merge").strip().lower()
        if "merge" not in action:
            continue
        title = str(r.get("Title") or "").strip()
        artist = str(r.get("Artist(s)") or "").strip()
        chart_type = str(r.get("Type") or "").strip()
        if title and artist and chart_type:
            broad_key = (_broad_normalize(title), _broad_normalize(artist), chart_type)
            existing = pairs.get(broad_key)
            if existing and existing != (title, artist):
                # Two rows in Apply Merges share the same broad key but have
                # different canonical forms — keep the first and warn so the
                # workbook can be corrected rather than silently dropping one.
                import warnings
                warnings.warn(
                    f"Merge workbook collision: broad key {broad_key!r} maps to both "
                    f"{existing!r} and {(title, artist)!r}. Keeping the first.",
                    stacklevel=2,
                )
            else:
                pairs[broad_key] = (title, artist)
    return pairs


def apply_merge_canonicalization(data, merge_pairs):
    """
    Re-canonicalize titles and primary-artist credits for every entry whose
    broad-normalized (title, artist, chart_type) matches a pair in the merge
    workbook.

    This ensures the same song uses one consistent title+artist string across
    all months and platforms, which in turn causes import_master_workbook() to
    create a single Release record for it instead of separate duplicates.
    """
    if not merge_pairs:
        return data

    def _remap_combined(row, chart_type):
        key = (
            _broad_normalize(row["Title"]),
            _broad_normalize(str(row.get("Primary_Artist") or "")),
            chart_type,
        )
        if key in merge_pairs:
            canonical_title, canonical_artist = merge_pairs[key]
            row = dict(row)
            row["Title"] = canonical_title
            row["Primary_Artist"] = canonical_artist
        return row

    def _remap_platform(row, chart_type):
        primary = str(row.get("Primary_Artist") or row.get("Artist") or "")
        key = (_broad_normalize(row["Title"]), _broad_normalize(primary), chart_type)
        if key in merge_pairs:
            canonical_title, canonical_artist = merge_pairs[key]
            row = dict(row)
            row["Title"] = canonical_title
            if "Primary_Artist" in row:
                row["Primary_Artist"] = canonical_artist
            else:
                row["Artist"] = canonical_artist
        return row

    for chart_type in ("singles", "albums"):
        data[chart_type]["combined"] = [
            _remap_combined(row, chart_type) for row in data[chart_type]["combined"]
        ]
        data[chart_type]["platforms"] = [
            _remap_platform(row, chart_type) for row in data[chart_type]["platforms"]
        ]
    return data


MONTHS = [
    "September 2025",
    "October 2025",
    "November 2025",
    "December 2025",
    "January 2026",
    "February 2026",
    "March 2026",
    "April 2026",
    "May 2026",
    "June 2026",
]

COMBINED_HEADERS = [
    "Month",
    "Rank",
    "Title",
    "Primary_Artist",
    "Featured_Artists",
    "Combined_Points_Raw",
    "Display_Points",
    "Platforms",
    "Platforms_Max",
    "Weeks",
    "Release_Year",
    "Confidence",
]

PLATFORM_HEADERS = ["Month", "Platform", "Rank", "Title", "Artist", "Points", "Weeks"]

PLATFORM_DATA = [
    ("Apple Music", "apple-music", "#FC3C44", 100, 101),
    ("Audiomack", "audiomack", "#F68B1F", 100, 101),
    ("Boomplay", "boomplay", "#00FFFF", 100, 101),
    ("Spotify", "spotify", "#1DB954", 100, 101),
    ("YouTube", "youtube", "#FF0000", 100, 101),
    ("Shazam", "shazam", "#0088FF", 100, 101),
]

ARTIST_SEPARATOR_RE = re.compile(
    r"\s*(?:&|,|\bfeat(?:uring)?\.?(?=\s)|\bft\.?(?=\s)|\band\b)\s*",
    re.IGNORECASE,
)
TITLE_ALIASES = {
    "chai ya saa umi": "chai ya saa kumi",
    "hallelujah(zimbabwe gospel song 2026)": "hallelujah",
}
CONFIDENCE_PRIORITY = {
    "": 0,
    "estimated (verify)": 1,
    "feat. from source": 2,
    "web-verified": 3,
}
ARTIST_NAME_ALIASES = {
    "ali kiba": "Alikiba",
    "bnxn": "BNXN",
    "bruce africa": "Bruce Africa",
    "dj smallz 732": "DJ Smallz",
    "d voice": "D Voice",
    "dyana cods": "Dyana Cods",
    "h art the band": "H_art the Band",
    "jay melody": "Jay Melody",
    "juma jux": "Jux",
    "kusslove tz": "Kusslove",
    "lilmaina": "Lil Maina",
    "mboso": "Mbosso",
    "mejja genge": "Mejja",
}


def _clean_text(value):
    return " ".join(unicodedata.normalize("NFKC", str(value or "")).strip().split())


def _normalize_title(value):
    title = _clean_text(value).replace("’", "'").casefold()
    return TITLE_ALIASES.get(title, title)


def _raw_artist_key(value):
    name = _clean_text(value).casefold().replace("_", " ")
    name = re.sub(r"[^\w]+", " ", name, flags=re.UNICODE)
    return " ".join(name.split())


def _canonical_artist_name(value):
    name = _clean_text(value)
    return ARTIST_NAME_ALIASES.get(_raw_artist_key(name), name)


def _artist_key(value):
    return _raw_artist_key(_canonical_artist_name(value))


def _split_artist_credit(value):
    return [
        part
        for part in (_canonical_artist_name(item) for item in ARTIST_SEPARATOR_RE.split(_clean_text(value)))
        if part
    ]


def _featured_artists(value):
    return _split_artist_credit(value)


def _credit_members(primary_artist, featured_artists=""):
    members = [_canonical_artist_name(primary_artist), *_featured_artists(featured_artists)]
    seen = set()
    result = []
    for member in members:
        key = _artist_key(member)
        if key and key not in seen:
            result.append(member)
            seen.add(key)
    return result


def _format_featured_artists(members):
    return ", ".join(members)


def _format_artist_credit(primary_artist, featured_artists=""):
    members = _credit_members(primary_artist, featured_artists)
    if len(members) <= 1:
        return members[0] if members else ""
    if len(members) == 2:
        return " & ".join(members)
    return f"{', '.join(members[:-1])} & {members[-1]}"


def _release_key(title, primary_artist, featured_artists=""):
    return _normalize_title(title), _artist_key(primary_artist)


def _merge_featured_artists(primary_artist, *featured_values):
    primary_key = _artist_key(primary_artist)
    seen = set()
    merged = []
    for value in featured_values:
        for member in _featured_artists(value):
            key = _artist_key(member)
            if key and key != primary_key and key not in seen:
                merged.append(member)
                seen.add(key)
    return _format_featured_artists(merged)


def _canonicalize_master_data(data):
    catalog_by_type = {}

    for chart_type in ("singles", "albums"):
        catalog = {}
        by_title = defaultdict(list)
        for row in data[chart_type]["combined"]:
            primary_artist = _canonical_artist_name(row["Primary_Artist"])
            featured = _format_featured_artists(_featured_artists(row["Featured_Artists"]))
            key = _release_key(row["Title"], primary_artist, featured)
            confidence = _clean_text(row["Confidence"])
            current = catalog.get(key)
            candidate = {
                "Title": _clean_text(row["Title"]),
                "Primary_Artist": primary_artist,
                "Featured_Artists": featured,
                "Release_Year": row["Release_Year"],
                "Confidence": confidence,
                "member_keys": {_artist_key(member) for member in _credit_members(primary_artist, featured)},
            }
            if current is None:
                catalog[key] = candidate
            else:
                merged_featured = _merge_featured_artists(
                    current["Primary_Artist"], current["Featured_Artists"], candidate["Featured_Artists"]
                )
                current["Featured_Artists"] = merged_featured
                current["member_keys"] = {
                    _artist_key(member) for member in _credit_members(current["Primary_Artist"], merged_featured)
                }
                if CONFIDENCE_PRIORITY.get(confidence, 0) > CONFIDENCE_PRIORITY.get(current["Confidence"], 0):
                    current["Title"] = candidate["Title"]
                    current["Primary_Artist"] = candidate["Primary_Artist"]
                    current["Release_Year"] = candidate["Release_Year"]
                    current["Confidence"] = candidate["Confidence"]

        for key, record in catalog.items():
            by_title[key[0]].append(record)
        catalog_by_type[chart_type] = (catalog, by_title)

    normalized = {}
    for chart_type in ("singles", "albums"):
        _catalog, by_title = catalog_by_type[chart_type]
        grouped_platform_rows = defaultdict(dict)

        for row in data[chart_type]["platforms"]:
            title_key = _normalize_title(row["Title"])
            candidates = by_title.get(title_key, [])
            raw_members = _split_artist_credit(row["Artist"])
            raw_member_keys = {_artist_key(member) for member in raw_members if _artist_key(member)}

            match = None
            exact = [candidate for candidate in candidates if candidate["member_keys"] == raw_member_keys]
            if len(exact) == 1:
                match = exact[0]
            elif len(candidates) == 1 and (
                not raw_member_keys or candidates[0]["member_keys"] & raw_member_keys
            ):
                match = candidates[0]
            else:
                subset = [candidate for candidate in candidates if raw_member_keys and raw_member_keys.issubset(candidate["member_keys"])]
                if len(subset) == 1:
                    match = subset[0]
                elif candidates and raw_member_keys:
                    match = max(candidates, key=lambda candidate: len(candidate["member_keys"] & raw_member_keys))
                    if not (match["member_keys"] & raw_member_keys):
                        match = None

            if match:
                title = match["Title"]
                primary_artist = match["Primary_Artist"]
                featured = _merge_featured_artists(
                    primary_artist,
                    match["Featured_Artists"],
                    _format_featured_artists(raw_members),
                )
                release_year = match["Release_Year"]
                confidence = match["Confidence"]
            else:
                title = _clean_text(row["Title"])
                primary_artist = raw_members[0] if raw_members else "Unknown Artist"
                featured = _format_featured_artists(raw_members[1:])
                release_year = None
                confidence = "estimated (verify)"

            release_key = _release_key(title, primary_artist, featured)
            group_key = (row["Month"], row["Platform"], release_key)
            existing = grouped_platform_rows[group_key].get("row")
            if existing is None:
                grouped_platform_rows[group_key]["row"] = {
                    "Month": row["Month"],
                    "Platform": row["Platform"],
                    "Rank": int(row["Rank"]),
                    "Title": title,
                    "Artist": _format_artist_credit(primary_artist, featured),
                    "Primary_Artist": primary_artist,
                    "Featured_Artists": featured,
                    "Points": int(row["Points"]),
                    "Weeks": int(row["Weeks"]),
                    "Release_Year": release_year,
                    "Confidence": confidence,
                    "Source_Rank": int(row["Rank"]),
                }
            else:
                # Duplicate entry for the same song in the same month+platform.
                # A song can only occupy one rank on a chart — keep the better
                # rank (lower number = higher points) and discard the other.
                # Summing would artificially inflate the score for a data error.
                incoming_rank = int(row["Rank"])
                if incoming_rank < existing["Source_Rank"]:
                    existing["Points"] = int(row["Points"])
                    existing["Rank"] = incoming_rank
                    existing["Source_Rank"] = incoming_rank
                existing["Weeks"] = max(existing["Weeks"], int(row["Weeks"]))
                existing["Featured_Artists"] = _merge_featured_artists(
                    existing["Primary_Artist"], existing["Featured_Artists"], featured
                )
                existing["Artist"] = _format_artist_credit(
                    existing["Primary_Artist"], existing["Featured_Artists"]
                )

        platform_rows = [value["row"] for value in grouped_platform_rows.values()]
        release_feature_map = {}
        for row in platform_rows:
            release_key = _release_key(row["Title"], row["Primary_Artist"], row["Featured_Artists"])
            release_feature_map[release_key] = _merge_featured_artists(
                row["Primary_Artist"],
                release_feature_map.get(release_key, ""),
                row["Featured_Artists"],
            )
        for row in platform_rows:
            release_key = _release_key(row["Title"], row["Primary_Artist"], row["Featured_Artists"])
            row["Featured_Artists"] = release_feature_map[release_key]
            row["Artist"] = _format_artist_credit(row["Primary_Artist"], row["Featured_Artists"])

        platform_groups = defaultdict(list)
        for row in platform_rows:
            platform_groups[(row["Month"], row["Platform"])].append(row)

        ranked_platform_rows = []
        for month_label in MONTHS:
            platform_names = ["Apple Music", "Audiomack"] if chart_type == "albums" else [item[0] for item in PLATFORM_DATA]
            for platform_name in platform_names:
                rows = platform_groups[(month_label, platform_name)]
                rows.sort(key=lambda item: (-item["Points"], item["Source_Rank"], item["Title"].casefold()))
                for rank, row in enumerate(rows, start=1):
                    row["Rank"] = rank
                    ranked_platform_rows.append(row)

        combined_groups = defaultdict(dict)
        for row in ranked_platform_rows:
            release_key = _release_key(row["Title"], row["Primary_Artist"], row["Featured_Artists"])
            group = combined_groups[row["Month"]].setdefault(
                release_key,
                {
                    "Title": row["Title"],
                    "Primary_Artist": row["Primary_Artist"],
                    "Featured_Artists": row["Featured_Artists"],
                    "Combined_Points_Raw": 0,
                    "platforms": set(),
                    "Weeks": 0,
                    "Release_Year": row["Release_Year"],
                    "Confidence": row["Confidence"],
                },
            )
            group["Combined_Points_Raw"] += row["Points"]
            group["platforms"].add(row["Platform"])
            group["Weeks"] = max(group["Weeks"], row["Weeks"])
            group["Featured_Artists"] = _merge_featured_artists(
                group["Primary_Artist"], group["Featured_Artists"], row["Featured_Artists"]
            )

        combined_rows = []
        platform_max = 2 if chart_type == "albums" else 6
        for month_label in MONTHS:
            rows = list(combined_groups[month_label].values())
            rows.sort(
                key=lambda item: (
                    -item["Combined_Points_Raw"],
                    -len(item["platforms"]),
                    item["Title"].casefold(),
                )
            )
            for rank, row in enumerate(rows[:50], start=1):
                combined_rows.append(
                    {
                        "Month": month_label,
                        "Rank": rank,
                        "Title": row["Title"],
                        "Primary_Artist": row["Primary_Artist"],
                        "Featured_Artists": row["Featured_Artists"],
                        "Combined_Points_Raw": row["Combined_Points_Raw"],
                        "Display_Points": 51 - rank,
                        "Platforms": len(row["platforms"]),
                        "Platforms_Max": platform_max,
                        "Weeks": row["Weeks"],
                        "Release_Year": row["Release_Year"],
                        "Confidence": row["Confidence"],
                    }
                )

        normalized[chart_type] = {
            "combined": combined_rows,
            "platforms": ranked_platform_rows,
        }

    return normalized


def _sheet_records(workbook, sheet_name, expected_headers):
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    if headers != expected_headers:
        raise ValueError(f"{sheet_name} headers do not match the documented workbook schema")
    return [dict(zip(headers, row)) for row in rows if any(value is not None for value in row)]


def load_master_workbook(workbook_path):
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    required = {
        "README",
        "Singles_Combined",
        "Albums_Combined",
        "Singles_Platforms",
        "Albums_Platforms",
    }
    if set(workbook.sheetnames) != required:
        raise ValueError("The master workbook must contain exactly the five documented sheets")

    raw_data = {
        "singles": {
            "combined": _sheet_records(workbook, "Singles_Combined", COMBINED_HEADERS),
            "platforms": _sheet_records(workbook, "Singles_Platforms", PLATFORM_HEADERS),
        },
        "albums": {
            "combined": _sheet_records(workbook, "Albums_Combined", COMBINED_HEADERS),
            "platforms": _sheet_records(workbook, "Albums_Platforms", PLATFORM_HEADERS),
        },
    }
    workbook.close()
    validate_master_data(raw_data)
    data = _canonicalize_master_data(raw_data)
    validate_master_data(data)
    return data


def validate_master_data(data):
    combined_total = 0
    for chart_type in ("singles", "albums"):
        rows = data[chart_type]["combined"]
        combined_total += len(rows)
        grouped = defaultdict(list)
        for row in rows:
            grouped[row["Month"]].append(row)

        if list(grouped) != MONTHS:
            raise ValueError(f"{chart_type} Combined months are missing or out of order")

        expected_max = 6 if chart_type == "singles" else 2
        for month in MONTHS:
            month_rows = grouped[month]
            ranks = [int(row["Rank"]) for row in month_rows]
            if len(month_rows) != 50 or ranks != list(range(1, 51)):
                raise ValueError(f"{chart_type} {month} must contain ranks 1 through 50 exactly once")
            for row in month_rows:
                rank = int(row["Rank"])
                if int(row["Display_Points"]) != 51 - rank:
                    raise ValueError(f"Invalid Display_Points at {chart_type} {month} rank {rank}")
                if int(row["Platforms_Max"]) != expected_max:
                    raise ValueError(f"Invalid Platforms_Max at {chart_type} {month} rank {rank}")

    if combined_total != 50 * 2 * len(MONTHS):
        raise ValueError(f"The two Combined sheets must contain exactly {50 * 2 * len(MONTHS)} rows")


def _month_parts(label):
    parsed = datetime.strptime(label, "%B %Y")
    return parsed.year, parsed.month


def _safe_slug(name, used):
    base = slugify(name)[:45] or "unknown"
    candidate = base
    suffix = 2
    while candidate in used and used[candidate] != name:
        candidate = f"{base[:40]}-{suffix}"
        suffix += 1
    used[candidate] = name
    return candidate


def import_master_workbook(app_registry, workbook_path, clear=True, write_line=None, merge_workbook_path=None):
    data = load_master_workbook(workbook_path)
    write_line = write_line or (lambda _message: None)

    # Apply merge-workbook canonicalization before any DB writes so duplicate
    # songs resolve to one canonical title+artist and are never inserted as
    # separate Release records in the first place.
    if merge_workbook_path:
        merge_path = Path(merge_workbook_path)
        if merge_path.exists():
            merge_pairs = load_merge_pairs(merge_path)
            if merge_pairs:
                data = apply_merge_canonicalization(data, merge_pairs)
                write_line(f"Merge canonicalization: applied {len(merge_pairs)} pairs from {merge_path.name}")

    Platform = app_registry.get_model("charts", "Platform")
    Artist = app_registry.get_model("charts", "Artist")
    Release = app_registry.get_model("charts", "Release")
    MonthlyChart = app_registry.get_model("charts", "MonthlyChart")
    MonthlyChartEntry = app_registry.get_model("charts", "MonthlyChartEntry")
    WeeklyUpload = app_registry.get_model("charts", "WeeklyUpload")
    Certification = app_registry.get_model("charts", "Certification")

    for name, slug, color, chart_size, points_base in PLATFORM_DATA:
        Platform.objects.update_or_create(
            name=name,
            defaults={
                "slug": slug,
                "color": color,
                "chart_size": chart_size,
                "points_base": points_base,
                "active": True,
            },
        )
    platforms = {platform.name: platform for platform in Platform.objects.all()}

    source_artist_names = {
        str(row["Primary_Artist"]).strip()
        for chart_type in data.values()
        for row in chart_type["combined"]
    }
    source_artist_names.update(
        str(row.get("Primary_Artist") or row["Artist"]).strip()
        for chart_type in data.values()
        for row in chart_type["platforms"]
    )

    if clear:
        write_line("Clearing existing singles and albums chart data...")
        Certification.objects.all().delete()
        MonthlyChartEntry.objects.all().delete()
        MonthlyChart.objects.all().delete()
        WeeklyUpload.objects.all().delete()
        Release.objects.all().delete()
        Artist.objects.exclude(name__in=source_artist_names).delete()

    artist_cache = {artist.name: artist for artist in Artist.objects.all()}
    used_slugs = {artist.slug: artist.name for artist in artist_cache.values()}
    release_cache = {}
    chart_cache = {}

    def get_artist(name):
        name = str(name).strip()
        artist = artist_cache.get(name)
        metadata = artist_country(name)
        if artist is None:
            defaults = {"name": name, "slug": _safe_slug(name, used_slugs)}
            if metadata:
                defaults.update(country=metadata[0], country_code=metadata[1])
            artist = Artist.objects.create(**defaults)
            artist_cache[name] = artist
        elif metadata and (not artist.country or not artist.country_code):
            artist.country, artist.country_code = metadata
            artist.save(update_fields=["country", "country_code"])
        return artist

    def get_release(title, artist_name, chart_type):
        title = str(title).strip()
        artist = get_artist(artist_name)
        canonical_title = _normalize_title(title)
        key = (chart_type, canonical_title, artist.pk)
        release = release_cache.get(key)
        if release is None:
            release = Release.objects.create(
                title=title,
                artist=artist,
                chart_type=chart_type,
                canonical_title=canonical_title,
            )
            release_cache[key] = release
        return release

    def get_chart(chart_type, month_label):
        key = (chart_type, month_label)
        chart = chart_cache.get(key)
        if chart is None:
            year, month_number = _month_parts(month_label)
            chart = MonthlyChart.objects.create(
                year=year,
                month=month_number,
                chart_type=chart_type,
                label=month_label,
                is_published=True,
            )
            chart_cache[key] = chart
        return chart

    combined_count = 0
    platform_count = 0

    for chart_type in ("singles", "albums"):
        combined_groups = defaultdict(list)
        for row in data[chart_type]["combined"]:
            combined_groups[row["Month"]].append(row)

        previous_rank = {}
        peak_rank = {}
        for month_label in MONTHS:
            chart = get_chart(chart_type, month_label)
            rows = combined_groups[month_label]
            current_best = {}
            for row in rows:
                key = (str(row["Title"]).strip().casefold(), str(row["Primary_Artist"]).strip().casefold())
                current_best[key] = min(current_best.get(key, 999), int(row["Rank"]))

            entries = []
            for row in rows:
                key = (str(row["Title"]).strip().casefold(), str(row["Primary_Artist"]).strip().casefold())
                rank = int(row["Rank"])
                release = get_release(row["Title"], row["Primary_Artist"], chart_type)
                entries.append(
                    MonthlyChartEntry(
                        chart=chart,
                        platform=None,
                        release=release,
                        rank=rank,
                        total_points=int(row["Display_Points"]),
                        raw_total_points=int(row["Combined_Points_Raw"]),
                        weeks_on_chart=int(row["Weeks"]),
                        platform_count=int(row["Platforms"]),
                        platform_max=int(row["Platforms_Max"]),
                        featured_artists=str(row["Featured_Artists"] or "").strip(),
                        release_year=int(row["Release_Year"]) if row["Release_Year"] is not None else None,
                        confidence=str(row["Confidence"] or "").strip(),
                        peak_rank=min(peak_rank.get(key, 999), current_best[key]),
                        prev_rank=previous_rank.get(key),
                    )
                )
            MonthlyChartEntry.objects.bulk_create(entries, batch_size=500)
            combined_count += len(entries)
            for key, rank in current_best.items():
                peak_rank[key] = min(peak_rank.get(key, 999), rank)
            previous_rank = current_best

        platform_groups = defaultdict(list)
        for row in data[chart_type]["platforms"]:
            platform_groups[(row["Month"], row["Platform"])].append(row)

        platform_previous = {}
        platform_peak = {}
        platform_names = ["Apple Music", "Audiomack"] if chart_type == "albums" else [item[0] for item in PLATFORM_DATA]
        for month_label in MONTHS:
            chart = get_chart(chart_type, month_label)
            for platform_name in platform_names:
                platform = platforms[platform_name]
                rows = platform_groups[(month_label, platform_name)]
                entries = []
                current_top_50 = {}
                for row in rows:
                    key = (
                        platform_name,
                        _normalize_title(row["Title"]),
                        _artist_key(row.get("Primary_Artist") or row["Artist"]),
                    )
                    rank = int(row["Rank"])
                    primary_artist = row.get("Primary_Artist") or _split_artist_credit(row["Artist"])[0]
                    featured_artists = str(row.get("Featured_Artists") or "").strip()
                    release = get_release(row["Title"], primary_artist, chart_type)
                    entries.append(
                        MonthlyChartEntry(
                            chart=chart,
                            platform=platform,
                            release=release,
                            rank=rank,
                            total_points=public_points(rank),
                            raw_total_points=max(int(row["Points"]), 0),
                            weeks_on_chart=int(row["Weeks"]),
                            platform_count=1,
                            platform_max=1,
                            featured_artists=featured_artists,
                            release_year=int(row["Release_Year"]) if row.get("Release_Year") is not None else None,
                            confidence=str(row.get("Confidence") or "").strip(),
                            peak_rank=min(platform_peak.get(key, 999), rank),
                            prev_rank=platform_previous.get(key),
                        )
                    )
                    if rank <= 50:
                        current_top_50[key] = rank
                        platform_peak[key] = min(platform_peak.get(key, 999), rank)
                MonthlyChartEntry.objects.bulk_create(entries, batch_size=500)
                platform_count += len(entries)
                platform_previous = {
                    key: value for key, value in platform_previous.items() if key[0] != platform_name
                }
                platform_previous.update(current_top_50)

    certification_rows = []
    cumulative_points = (
        MonthlyChartEntry.objects.filter(platform__isnull=True)
        .values("release_id")
        .annotate(total=Sum("total_points"))
    )
    for item in cumulative_points:
        total = int(item["total"] or 0)
        for level, threshold in CERTIFICATION_THRESHOLDS.items():
            if total >= threshold:
                certification_rows.append(
                    Certification(
                        release_id=item["release_id"],
                        level=level,
                        total_points=total,
                    )
                )
    Certification.objects.bulk_create(certification_rows, batch_size=500)

    write_line(
        f"Imported {combined_count} Combined rows, {platform_count} platform rows, "
        f"and {len(certification_rows)} certifications"
    )
    return {
        "combined_rows": combined_count,
        "platform_rows": platform_count,
        "total_rows": combined_count + platform_count,
        "certifications": len(certification_rows),
        "months": MONTHS,
    }
