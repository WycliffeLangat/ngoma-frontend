"""
Merge approved duplicate releases from the Ngoma duplicate-release workbook.

Usage:
    python manage.py merge_duplicate_releases --file "Data/ngoma_duplicate_releases_final_merge_ready.xlsx" --dry-run
    python manage.py merge_duplicate_releases --file "Data/ngoma_duplicate_releases_final_merge_ready.xlsx" --apply
"""

from collections import defaultdict
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

import openpyxl

from charts.models import (
    AuditLog,
    Certification,
    MonthlyChartEntry,
    PlatformChartEntry,
    Release,
    ReleaseArtistCredit,
)
from charts.cms_utils import recalculate_certifications, record_merge_normalization


# Fields copied from duplicate to keeper only when the keeper's field is blank/null/zero.
_METADATA_FIELDS = [
    "cover_image",
    "genre",
    "label",
    "distributor",
    "isrc",
    "upc",
    "release_year",
    "release_date",
    "featured_artists",
    "credited_artists",
    "songwriters",
    "producers",
    "spotify_url",
    "apple_music_url",
    "youtube_url",
    "boomplay_url",
    "audiomack_url",
    "tiktok_url",
    "shazam_url",
    "radio_info",
]


def _sheet_rows(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        return [], []
    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    data = [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]
    return headers, data


def _to_int(value):
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _load_merge_plan(filepath):
    """Return (merge_pairs, protected_ids) from the workbook."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    _, apply_rows = _sheet_rows(wb, "Apply Merges")
    _, keep_rows = _sheet_rows(wb, "Keep Separate")
    wb.close()

    protected_ids = set()
    for row in keep_rows:
        rid = _to_int(row.get("Release ID"))
        if rid:
            protected_ids.add(rid)

    merge_pairs = []
    for row in apply_rows:
        action = str(row.get("Action") or "").strip().lower()
        if "merge" not in action:
            continue
        dup_id = _to_int(row.get("Duplicate Release ID"))
        keep_id = _to_int(row.get("Keep Release ID"))
        if dup_id and keep_id and dup_id != keep_id:
            merge_pairs.append({
                "dup_id": dup_id,
                "keep_id": keep_id,
                "title": str(row.get("Title") or "").strip(),
                "artist": str(row.get("Artist(s)") or "").strip(),
                "chart_type": str(row.get("Type") or "").strip(),
                "reason": str(row.get("Why merge") or "").strip(),
            })

    return merge_pairs, protected_ids


def _find_release_by_title(title, artist_str, chart_type):
    """
    Fallback lookup when a release ID is not found.
    Tries canonical_title + each artist name listed in artist_str.
    """
    canonical = title.strip().lower()
    # artist_str may contain multiple names separated by commas, &, or |
    import re as _re
    names = [n.strip() for n in _re.split(r"[,&|]", artist_str) if n.strip()]
    for name in names:
        from django.db.models import Q
        artist = (
            Release.objects.filter(
                canonical_title=canonical,
                chart_type=chart_type,
                artist__name__iexact=name,
            ).first()
        )
        if artist:
            return artist
    return None


def _resolve_release(release_id, title, artist_str, chart_type):
    """
    Try to find a Release by primary key first.
    Falls back to title+artist+chart_type when the ID is absent from the DB
    (production DB IDs may differ from the development DB used to build the workbook).
    """
    try:
        return Release.objects.get(pk=release_id)
    except Release.DoesNotExist:
        pass
    return _find_release_by_title(title, artist_str, chart_type)


def _merge_chart_entries(dup, keeper, dry_run):
    """
    Reassign MonthlyChartEntry records from dup to keeper.

    Where keeper already has an entry in the same (chart, platform), the two
    entries are consolidated: points are summed, the keeper's rank/position is
    kept, duplicate entry is deleted.  Where there is no conflict the entry's
    release FK is simply updated to keeper.

    Returns (reassigned_count, merged_count).
    """
    reassigned = 0
    merged = 0

    dup_entries = list(MonthlyChartEntry.objects.filter(release=dup).select_related("chart"))
    for dup_entry in dup_entries:
        keeper_entry = MonthlyChartEntry.objects.filter(
            chart_id=dup_entry.chart_id,
            platform_id=dup_entry.platform_id,
            release=keeper,
        ).first()

        if keeper_entry:
            merged += 1
            if not dry_run:
                keeper_entry.total_points += dup_entry.total_points
                if dup_entry.raw_total_points is not None:
                    keeper_entry.raw_total_points = (
                        (keeper_entry.raw_total_points or 0) + dup_entry.raw_total_points
                    )
                keeper_entry.weeks_on_chart = max(
                    keeper_entry.weeks_on_chart, dup_entry.weeks_on_chart
                )
                keeper_entry.peak_rank = min(
                    keeper_entry.peak_rank or 9999, dup_entry.peak_rank or 9999
                )
                if dup_entry.platform_id is None:
                    # Combined entry: platform_count tracks how many platforms charted.
                    keeper_entry.platform_count = min(
                        keeper_entry.platform_max,
                        keeper_entry.platform_count + dup_entry.platform_count,
                    )
                keeper_entry.save()
                dup_entry.delete()
        else:
            reassigned += 1
            if not dry_run:
                dup_entry.release = keeper
                dup_entry.save(update_fields=["release"])

    return reassigned, merged


def _merge_platform_chart_entries(dup, keeper, dry_run):
    """Reassign PlatformChartEntry records from dup to keeper (simple FK update)."""
    count = PlatformChartEntry.objects.filter(release=dup).count()
    if not dry_run and count:
        PlatformChartEntry.objects.filter(release=dup).update(release=keeper)
    return count


def _merge_artist_credits(dup, keeper, dry_run):
    """
    Move unique ReleaseArtistCredit records from dup to keeper.

    Credits already present on keeper (same artist + role) are skipped.
    Returns number of credits moved.
    """
    existing = set(
        ReleaseArtistCredit.objects.filter(release=keeper).values_list("artist_id", "role")
    )
    moved = 0
    for credit in ReleaseArtistCredit.objects.filter(release=dup):
        if (credit.artist_id, credit.role) not in existing:
            moved += 1
            if not dry_run:
                max_pos = ReleaseArtistCredit.objects.filter(
                    release=keeper, role=credit.role
                ).count()
                ReleaseArtistCredit.objects.create(
                    release=keeper,
                    artist_id=credit.artist_id,
                    role=credit.role,
                    position=max_pos,
                )
                existing.add((credit.artist_id, credit.role))
    return moved


def _merge_metadata(dup, keeper, dry_run):
    """Copy metadata from dup to keeper for any field that is blank on keeper."""
    filled = []
    for field in _METADATA_FIELDS:
        dup_val = getattr(dup, field, None)
        keeper_val = getattr(keeper, field, None)
        if dup_val and not keeper_val:
            filled.append(field)
            if not dry_run:
                setattr(keeper, field, dup_val)
    if filled and not dry_run:
        keeper.save(update_fields=filled + ["updated_at"])
    return filled


def _merge_certifications(dup, keeper, dry_run):
    """Delete dup's certifications — recalculate_certifications handles keeper's."""
    count = Certification.objects.filter(release=dup).count()
    if not dry_run and count:
        Certification.objects.filter(release=dup).delete()
    return count


def _archive_duplicate(dup, keeper, dry_run):
    if not dry_run:
        dup.status = "archived"
        dup.save(update_fields=["status", "updated_at"])
        record_merge_normalization(
            "title", dup.title, keeper.title,
            notes=f"Auto-created when release {dup.pk} was merged into {keeper.pk}",
        )


def _log_merge(dup, keeper, stats):
    try:
        AuditLog.objects.create(
            action="merge_release",
            module="releases",
            object_type="Release",
            object_id=str(keeper.pk),
            object_repr=str(keeper)[:255],
            old_value={"merged_release_id": dup.pk, "merged_release": str(dup)[:255]},
            new_value=stats,
            reason="Automated duplicate-release merge from approved workbook",
            user=None,
            ip_address=None,
            user_agent="management-command",
        )
    except Exception:
        pass


class Command(BaseCommand):
    help = (
        "Merge approved duplicate releases from the Ngoma duplicate-release workbook. "
        "Pass --dry-run to preview changes or --apply to execute them."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Path to the merge workbook (e.g. Data/ngoma_duplicate_releases_final_merge_ready.xlsx)",
        )
        group = parser.add_mutually_exclusive_group(required=False)
        group.add_argument("--dry-run", action="store_true", help="Preview changes without modifying the database")
        group.add_argument("--apply", action="store_true", help="Apply all approved merges (default)")

    def handle(self, *args, **options):
        filepath = Path(options["file"])
        if not filepath.exists():
            raise CommandError(f"File not found: {filepath}")

        dry_run = options["dry_run"]  # --apply is the default if neither flag is given
        mode_label = "DRY RUN" if dry_run else "APPLY"
        self.stdout.write(self.style.WARNING(f"\n[{mode_label}] Loading merge plan from: {filepath}\n"))

        merge_pairs, protected_ids = _load_merge_plan(filepath)
        self.stdout.write(f"  Merge pairs loaded : {len(merge_pairs)}")
        self.stdout.write(f"  Protected IDs      : {len(protected_ids)}\n")

        # Validate and filter pairs
        valid_pairs = []
        skipped = []
        for pair in merge_pairs:
            dup_id, keep_id = pair["dup_id"], pair["keep_id"]
            issues = []
            if dup_id in protected_ids:
                issues.append(f"dup_id {dup_id} is in Keep Separate list")
            if keep_id in protected_ids:
                issues.append(f"keep_id {keep_id} is in Keep Separate list")
            if issues:
                skipped.append((pair, issues))
                continue

            dup = _resolve_release(dup_id, pair["title"], pair["artist"], pair["chart_type"])
            if dup is None:
                skipped.append((pair, [f"Duplicate release {dup_id} not found by ID or title+artist"]))
                continue

            keeper = _resolve_release(keep_id, pair["title"], pair["artist"], pair["chart_type"])
            if keeper is None:
                skipped.append((pair, [f"Keeper release {keep_id} not found by ID or title+artist"]))
                continue

            if dup.status == "archived":
                skipped.append((pair, [f"Duplicate release {dup_id} is already archived"]))
                continue

            valid_pairs.append((pair, dup, keeper))

        self.stdout.write(f"  Valid pairs to process : {len(valid_pairs)}")
        self.stdout.write(f"  Pairs skipped          : {len(skipped)}\n")

        def _s(text):
            """ASCII-safe string for Windows console output."""
            return str(text).encode("ascii", "replace").decode("ascii")

        if skipped:
            self.stdout.write(self.style.WARNING("Skipped pairs:"))
            for pair, reasons in skipped:
                self.stdout.write(
                    f"  [{pair['dup_id']} -> {pair['keep_id']}] {_s(pair['title'])!r}: {'; '.join(reasons)}"
                )
            self.stdout.write("")

        if not valid_pairs:
            self.stdout.write(self.style.SUCCESS("Nothing to merge."))
            return

        # Execute merges
        total_mce_reassigned = 0
        total_mce_merged = 0
        total_pce = 0
        total_rac = 0
        affected_keeper_ids = set()

        def _run_merges():
            nonlocal total_mce_reassigned, total_mce_merged, total_pce, total_rac

            for idx, (pair, dup, keeper) in enumerate(valid_pairs, 1):
                self.stdout.write(
                    f"  [{idx:>3}/{len(valid_pairs)}] {pair['dup_id']} -> {pair['keep_id']}  "
                    f"{_s(pair['title'])!r} / {_s(pair['artist'])!r}"
                )

                mce_reassigned, mce_merged = _merge_chart_entries(dup, keeper, dry_run)
                pce_count = _merge_platform_chart_entries(dup, keeper, dry_run)
                rac_count = _merge_artist_credits(dup, keeper, dry_run)
                meta_filled = _merge_metadata(dup, keeper, dry_run)
                cert_deleted = _merge_certifications(dup, keeper, dry_run)
                _archive_duplicate(dup, keeper, dry_run)

                total_mce_reassigned += mce_reassigned
                total_mce_merged += mce_merged
                total_pce += pce_count
                total_rac += rac_count
                affected_keeper_ids.add(keeper.pk)

                details = []
                if mce_reassigned:
                    details.append(f"mce_reassigned={mce_reassigned}")
                if mce_merged:
                    details.append(f"mce_merged={mce_merged}")
                if pce_count:
                    details.append(f"pce={pce_count}")
                if rac_count:
                    details.append(f"rac_moved={rac_count}")
                if meta_filled:
                    details.append(f"meta={meta_filled}")
                if cert_deleted:
                    details.append(f"certs_deleted={cert_deleted}")

                if details:
                    self.stdout.write(f"       -> {', '.join(details)}")

                if not dry_run:
                    _log_merge(dup, keeper, {
                        "mce_reassigned": mce_reassigned,
                        "mce_merged": mce_merged,
                        "pce_reassigned": pce_count,
                        "rac_moved": rac_count,
                        "meta_filled": meta_filled,
                        "certs_deleted": cert_deleted,
                    })

        if dry_run:
            _run_merges()
        else:
            with transaction.atomic():
                _run_merges()

        # Summary
        self.stdout.write("")
        self.stdout.write("=" * 60)
        self.stdout.write(f"  Pairs processed            : {len(valid_pairs)}")
        self.stdout.write(f"  Chart entries reassigned   : {total_mce_reassigned}")
        self.stdout.write(f"  Chart entries merged       : {total_mce_merged}")
        self.stdout.write(f"  Platform entries moved     : {total_pce}")
        self.stdout.write(f"  Artist credits moved       : {total_rac}")
        self.stdout.write(f"  Affected keeper releases   : {len(affected_keeper_ids)}")
        self.stdout.write("=" * 60)

        if not dry_run:
            self.stdout.write("\nRecalculating certifications for all affected releases...")
            updated = 0
            for release_id in affected_keeper_ids:
                try:
                    release = Release.objects.get(pk=release_id)
                    updated += recalculate_certifications(release=release)
                except Release.DoesNotExist:
                    pass
            self.stdout.write(f"  Certifications updated: {updated}")

            # Create a single AuditLog entry for the overall operation so the
            # public-data revision bumps and the frontend cache invalidates.
            try:
                AuditLog.objects.create(
                    action="bulk_merge_releases",
                    module="releases",
                    object_type="Release",
                    object_id="",
                    object_repr=f"Merged {len(valid_pairs)} duplicate releases",
                    old_value={},
                    new_value={
                        "pairs_merged": len(valid_pairs),
                        "pairs_skipped": len(skipped),
                        "mce_reassigned": total_mce_reassigned,
                        "mce_merged": total_mce_merged,
                        "certifications_updated": updated,
                        "timestamp": timezone.now().isoformat(),
                    },
                    reason="Automated bulk merge from approved workbook",
                    user=None,
                    ip_address=None,
                    user_agent="management-command",
                )
            except Exception:
                pass

            self.stdout.write(self.style.SUCCESS(
                f"\nDone. {len(valid_pairs)} duplicate releases merged and archived."
            ))
        else:
            self.stdout.write(self.style.WARNING(
                "\nDry run complete. No changes were made. Re-run with --apply to execute."
            ))
