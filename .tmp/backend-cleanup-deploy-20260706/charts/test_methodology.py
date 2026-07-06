import io

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase

from .artist_credits import parse_artist_credit
from .cms_utils import (
    harmonize_chart_history,
    published_top50_entries,
    recalculate_certifications,
)
from .models import (
    Artist,
    Certification,
    CertificationRule,
    ChartType,
    MonthlyChart,
    MonthlyChartEntry,
    Platform,
    PlatformChartEntry,
    Release,
    WeeklyUpload,
)
from .pipeline import process_weekly_upload, rebuild_monthly_chart


PLATFORM_NAMES = [
    "Apple Music",
    "Audiomack",
    "Boomplay",
    "Spotify",
    "YouTube",
    "Shazam",
]


def workbook_bytes(size):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Apple Music"])
    for rank in range(1, size + 1):
        sheet.append([f"Song {rank} - Artist {rank}"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class CanonicalMethodologyTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.platforms = {}
        for order, name in enumerate(PLATFORM_NAMES):
            cls.platforms[name], _ = Platform.objects.update_or_create(
                name=name,
                defaults={
                    "slug": name.lower().replace(" ", "-"),
                    "chart_size": 100,
                    "max_chart_size": 50,
                    "points_base": 999,  # mutable metadata cannot alter scoring
                    "supports_singles": True,
                    "supports_albums": name in {"Apple Music", "Audiomack"},
                    "display_order": order,
                },
            )

    def upload(self, year, month, week, chart_type=ChartType.SINGLES):
        return WeeklyUpload.objects.create(
            chart_type=chart_type,
            year=year,
            month=month,
            week=week,
            file=SimpleUploadedFile("weekly.xlsx", b"placeholder"),
        )

    def release(self, title, artist_name=None, chart_type=ChartType.SINGLES):
        artist_name = artist_name or f"{title} Artist"
        artist, _ = Artist.objects.get_or_create(
            name=artist_name,
            defaults={"slug": artist_name.lower().replace(" ", "-")},
        )
        release, _ = Release.objects.get_or_create(
            canonical_title=title.lower(),
            artist=artist,
            chart_type=chart_type,
            defaults={"title": title},
        )
        return release

    def test_weekly_top_100_cap_and_fixed_points(self):
        upload = self.upload(2026, 7, 1)
        process_weekly_upload(upload, io.BytesIO(workbook_bytes(105)))

        entries = PlatformChartEntry.objects.filter(upload=upload)
        self.assertEqual(entries.count(), 100)
        self.assertEqual(entries.get(position=1).points, 100)
        self.assertEqual(entries.get(position=100).points, 1)
        self.assertFalse(entries.filter(position__gt=100).exists())
        self.assertFalse(entries.filter(points__lte=0).exists())

    def test_weekly_short_chart_creates_no_blank_rows(self):
        upload = self.upload(2026, 8, 1)
        process_weekly_upload(upload, io.BytesIO(workbook_bytes(75)))

        entries = PlatformChartEntry.objects.filter(upload=upload)
        self.assertEqual(entries.count(), 75)
        self.assertEqual(entries.get(position=75).points, 26)
        self.assertFalse(entries.filter(position__gt=75).exists())

    def test_blank_weekly_row_stays_empty_without_shifting_later_ranks(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Apple Music"])
        sheet.append(["First - Artist A"])
        sheet.append([None])
        sheet.append(["Third - Artist C"])
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        upload = self.upload(2026, 12, 1)
        process_weekly_upload(upload, io.BytesIO(output.getvalue()))
        entries = PlatformChartEntry.objects.filter(upload=upload)
        self.assertEqual(list(entries.values_list("position", "points")), [(1, 100), (3, 98)])

    def test_albums_use_the_same_top_100_source_cap(self):
        upload = self.upload(2027, 1, 1, chart_type=ChartType.ALBUMS)
        process_weekly_upload(upload, io.BytesIO(workbook_bytes(105)))
        entries = PlatformChartEntry.objects.filter(upload=upload)
        self.assertEqual(entries.count(), 100)
        self.assertEqual(entries.get(position=100).points, 1)
        combined = MonthlyChartEntry.objects.get(
            chart__year=2027,
            chart__month=1,
            platform__isnull=True,
            rank=1,
        )
        self.assertEqual(combined.platform_max, 2)

    def test_monthly_platform_sum_and_combined_raw_score(self):
        release = self.release("Finale", "Bien")
        uploads = [self.upload(2026, 9, week) for week in (1, 2, 3)]
        positions = {
            "Apple Music": [1, 3, 50],       # 249
            "Spotify": [1, 1, 1],            # 300
            "YouTube": [31, 31, 31],         # 210
            "Boomplay": [41, 41, 41],        # 180
            "Audiomack": [31, None, None],   # 70
            "Shazam": [None, None, None],    # 0
        }
        for upload in uploads:
            upload.processed = True
            upload.save(update_fields=["processed"])
        for platform_name, ranks in positions.items():
            for upload, rank in zip(uploads, ranks):
                if rank is None:
                    continue
                PlatformChartEntry.objects.create(
                    upload=upload,
                    platform=self.platforms[platform_name],
                    release=release,
                    position=rank,
                    points=1,  # rebuild derives the canonical score from rank
                    raw_title=release.title,
                    raw_artist=release.artist.name,
                )

        rebuild_monthly_chart(ChartType.SINGLES, 2026, 9)
        apple = MonthlyChartEntry.objects.get(
            chart__year=2026,
            chart__month=9,
            platform=self.platforms["Apple Music"],
            release=release,
        )
        combined = MonthlyChartEntry.objects.get(
            chart__year=2026,
            chart__month=9,
            platform__isnull=True,
            release=release,
        )
        self.assertEqual(apple.raw_total_points, 249)
        self.assertEqual(apple.total_points, 50)
        self.assertEqual(combined.raw_total_points, 1009)
        self.assertEqual(combined.total_points, 50)
        self.assertEqual((combined.platform_count, combined.platform_max), (5, 6))

    def test_rank_50_gets_one_public_point_and_rank_51_gets_zero(self):
        upload = self.upload(2026, 10, 1)
        upload.processed = True
        upload.save(update_fields=["processed"])
        for rank in range(1, 52):
            release = self.release(f"Rank {rank}")
            PlatformChartEntry.objects.create(
                upload=upload,
                platform=self.platforms["Apple Music"],
                release=release,
                position=rank,
                points=101 - rank,
                raw_title=release.title,
                raw_artist=release.artist.name,
            )

        rebuild_monthly_chart(ChartType.SINGLES, 2026, 10)
        combined = MonthlyChartEntry.objects.filter(
            chart__year=2026,
            chart__month=10,
            platform__isnull=True,
        )
        self.assertEqual(combined.get(rank=50).total_points, 1)
        self.assertEqual(combined.get(rank=51).total_points, 0)

    def test_certifications_sum_only_published_combined_top_50(self):
        release = self.release("Certification Song")
        for month, rank, points in [(7, 1, 50), (8, 5, 46), (9, 54, 999)]:
            chart = MonthlyChart.objects.create(
                year=2027,
                month=month,
                chart_type=ChartType.SINGLES,
                status="published",
                is_published=True,
            )
            MonthlyChartEntry.objects.create(
                chart=chart,
                release=release,
                rank=rank,
                total_points=points,
                raw_total_points=points,
            )
        CertificationRule.objects.update_or_create(
            level="gold",
            defaults={"threshold": 96, "active": True},
        )
        CertificationRule.objects.filter(level__in=["platinum", "diamond"]).update(active=False)

        recalculate_certifications(release=release)

        certification = Certification.objects.get(release=release, level="gold")
        self.assertEqual(certification.total_points, 96)

    def test_collaborators_and_features_receive_structured_credits(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Apple Music"])
        sheet.append(["Finale - Bien & Alikiba"])
        sheet.append(["Lifestyle - Bien ft. Scar"])
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        upload = self.upload(2026, 11, 1)
        process_weekly_upload(upload, io.BytesIO(output.getvalue()))

        finale = Release.objects.get(
            canonical_title="finale",
            artist__name="Bien",
            chart_type=ChartType.SINGLES,
        )
        lifestyle = Release.objects.get(
            canonical_title="lifestyle",
            artist__name="Bien",
            chart_type=ChartType.SINGLES,
        )
        self.assertEqual(
            list(finale.artist_credits.order_by("position").values_list("artist__name", "role")),
            [("Bien", "primary"), ("Alikiba", "primary")],
        )
        self.assertEqual(
            set(lifestyle.artist_credits.values_list("artist__name", "role")),
            {("Bien", "primary"), ("Scar", "featured")},
        )

    def test_public_history_ignores_rank_54_and_marks_reentry(self):
        target = self.release("History Target")
        placements = [(7, 40), (8, 20), (9, 54), (10, 10)]
        for month, target_rank in placements:
            chart = MonthlyChart.objects.create(
                year=2028,
                month=month,
                chart_type=ChartType.SINGLES,
                status="published",
                is_published=True,
            )
            for rank in range(1, target_rank + 1):
                release = target if rank == target_rank else self.release(
                    f"Filler {month}-{rank}"
                )
                MonthlyChartEntry.objects.create(
                    chart=chart,
                    release=release,
                    rank=rank,
                    total_points=max(51 - rank, 0),
                    raw_total_points=1000 - rank,
                    platform_max=6,
                )

        harmonize_chart_history(chart_type=ChartType.SINGLES)
        october = MonthlyChartEntry.objects.get(
            chart__year=2028,
            chart__month=10,
            release=target,
            platform__isnull=True,
        )
        self.assertEqual(
            published_top50_entries().filter(release=target).count(),
            3,
        )
        self.assertEqual(october.prev_rank, None)
        self.assertEqual(october.peak_rank, 10)
        self.assertEqual(october.movement, "re-entry")

    def test_legitimate_group_name_is_not_split(self):
        self.assertEqual(parse_artist_credit("Years & Years"), (["Years & Years"], []))
