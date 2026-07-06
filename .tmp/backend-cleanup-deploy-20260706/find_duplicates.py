import sqlite3, re, unicodedata
from difflib import SequenceMatcher
from itertools import combinations
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

conn = sqlite3.connect('ngoma_charts.db')
cur = conn.cursor()
cur.execute("""
    SELECT
        r.id, r.title, r.chart_type, r.status, r.release_year,
        GROUP_CONCAT(a.name, ' | ') as artists,
        COUNT(mce.id) as chart_entries
    FROM charts_release r
    LEFT JOIN charts_releaseartistcredit rac ON rac.release_id = r.id AND rac.role = 'primary'
    LEFT JOIN charts_artist a ON a.id = rac.artist_id
    LEFT JOIN charts_monthlychartentry mce ON mce.release_id = r.id AND mce.platform_id IS NULL
    GROUP BY r.id
    ORDER BY r.title COLLATE NOCASE
""")
releases = cur.fetchall()
conn.close()
print(f"Loaded {len(releases)} releases")

def normalise(s):
    if not s:
        return ""
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\b(ft|feat|featuring|and|the|a|an|of|in|on|by)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def sim(a, b):
    return SequenceMatcher(None, a, b).ratio()

data = []
for row in releases:
    rid, title, chart_type, status, year, artists, entries = row
    data.append({
        'id': rid,
        'title': title or '',
        'chart_type': chart_type or '',
        'status': status or '',
        'year': year or '',
        'artists': artists or '',
        'entries': entries or 0,
        'norm_title': normalise(title or ''),
        'norm_artists': normalise(artists or ''),
    })

flagged = []
flagged_pairs = set()

# Pass 1: exact normalised title — all combinations within the group
by_norm_title = defaultdict(list)
for d in data:
    if d['norm_title']:
        by_norm_title[d['norm_title']].append(d)

for norm_title, group in by_norm_title.items():
    if len(group) < 2:
        continue
    for r1, r2 in combinations(group, 2):
        pair = (min(r1['id'], r2['id']), max(r1['id'], r2['id']))
        if pair in flagged_pairs:
            continue
        artist_sim = sim(r1['norm_artists'], r2['norm_artists']) if (r1['norm_artists'] or r2['norm_artists']) else 0
        parts = ["Identical title (normalised)"]
        if artist_sim > 0.7:
            parts.append(f"Similar artists ({int(artist_sim*100)}%)")
        elif r1['artists'] == r2['artists']:
            parts.append("Same artist")
        score = round(0.5 + 0.5 * artist_sim, 3)
        flagged.append({'reason': ' + '.join(parts), 'score': score, 'r1': r1, 'r2': r2})
        flagged_pairs.add(pair)

print(f"Pass 1 (exact normalised title): {len(flagged)} pairs")

# Pass 2: fuzzy title within same chart_type
TITLE_THRESHOLD = 0.82
by_type = defaultdict(list)
for d in data:
    by_type[d['chart_type']].append(d)

fuzzy_found = 0
for chart_type, group in by_type.items():
    print(f"  Fuzzy: {chart_type} ({len(group)} releases)...")
    for i in range(len(group)):
        for j in range(i+1, len(group)):
            r1, r2 = group[i], group[j]
            pair = (min(r1['id'], r2['id']), max(r1['id'], r2['id']))
            if pair in flagged_pairs:
                continue
            if not r1['norm_title'] or not r2['norm_title']:
                continue
            ts = sim(r1['norm_title'], r2['norm_title'])
            if ts < TITLE_THRESHOLD:
                continue
            as_ = sim(r1['norm_artists'], r2['norm_artists']) if (r1['norm_artists'] or r2['norm_artists']) else 0
            combined = round(0.6 * ts + 0.4 * as_, 3)
            if combined < 0.75:
                continue
            parts = [f"Similar title ({int(ts*100)}%)"]
            if as_ > 0.7:
                parts.append(f"Similar artists ({int(as_*100)}%)")
            flagged.append({'reason': ' + '.join(parts), 'score': combined, 'r1': r1, 'r2': r2})
            flagged_pairs.add(pair)
            fuzzy_found += 1

print(f"Pass 2 (fuzzy): {fuzzy_found} additional pairs")
print(f"Total pairs: {len(flagged)}")

flagged.sort(key=lambda x: -x['score'])

# Build Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Potential Duplicates"

header_fill = PatternFill("solid", fgColor="1A1A1A")
header_font = Font(bold=True, color="FFFFFF", size=10)
high_fill = PatternFill("solid", fgColor="FFD6D6")
mid_fill  = PatternFill("solid", fgColor="FFF3CD")
low_fill  = PatternFill("solid", fgColor="E8F5E9")
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = [
    "Pair #", "Reason", "Confidence",
    "ID A", "Title A", "Artists A", "Type A", "Entries A", "Status A",
    "ID B", "Title B", "Artists B", "Type B", "Entries B", "Status B",
    "Action"
]
ws.append(headers)
for col in range(1, len(headers)+1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

col_widths = [7, 42, 12, 8, 34, 34, 10, 9, 9, 8, 34, 34, 10, 9, 9, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30

for idx, f in enumerate(flagged, 1):
    r1, r2 = f['r1'], f['r2']
    score = f['score']
    if score >= 0.90:
        row_fill = high_fill
        conf = "HIGH"
    elif score >= 0.80:
        row_fill = mid_fill
        conf = "MEDIUM"
    else:
        row_fill = low_fill
        conf = "LOW"

    row_data = [
        idx, f['reason'], conf,
        r1['id'], r1['title'], r1['artists'], r1['chart_type'], r1['entries'], r1['status'],
        r2['id'], r2['title'], r2['artists'], r2['chart_type'], r2['entries'], r2['status'],
        ""
    ]
    ws.append(row_data)
    rn = ws.max_row
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=rn, column=col)
        cell.fill = row_fill
        cell.border = border
        wrap_cols = {2, 5, 6, 11, 12}
        center_cols = {1, 3, 4, 7, 8, 9, 10, 13, 14, 15}
        if col in wrap_cols:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        elif col in center_cols:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(vertical='center')
    ws.row_dimensions[rn].height = 18

ws.freeze_panes = "A2"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2.column_dimensions['A'].width = 38
ws2.column_dimensions['B'].width = 15
summary = [
    ("Total releases analysed", len(data)),
    ("Total potential duplicate pairs", len(flagged)),
    ("HIGH confidence (>=90%) - likely same release", sum(1 for f in flagged if f['score'] >= 0.90)),
    ("MEDIUM confidence (80-89%) - probable duplicate", sum(1 for f in flagged if 0.80 <= f['score'] < 0.90)),
    ("LOW confidence (75-79%) - needs review", sum(1 for f in flagged if f['score'] < 0.80)),
    ("", ""),
    ("Colour key:", ""),
    ("RED = HIGH confidence duplicate", ""),
    ("AMBER = MEDIUM confidence duplicate", ""),
    ("GREEN = LOW confidence / needs review", ""),
]
red_fill = PatternFill("solid", fgColor="FFD6D6")
amb_fill = PatternFill("solid", fgColor="FFF3CD")
grn_fill = PatternFill("solid", fgColor="E8F5E9")
for i, r in enumerate(summary, 1):
    ws2.append(list(r))
    ws2.cell(i, 1).font = Font(bold=True)
    ws2.cell(i, 2).alignment = Alignment(horizontal='right')
    if i == 8:
        ws2.cell(i, 1).fill = red_fill
    elif i == 9:
        ws2.cell(i, 1).fill = amb_fill
    elif i == 10:
        ws2.cell(i, 1).fill = grn_fill

out = r"c:\Users\HP\Desktop\Ngoma Charts Folder\files\duplicate_releases.xlsx"
wb.save(out)
print(f"\nSaved to: {out}")
