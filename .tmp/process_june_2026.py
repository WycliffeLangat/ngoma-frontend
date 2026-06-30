import argparse
import calendar
import io
import json
import os
import re
import sys
import unicodedata
import zipfile
from collections import defaultdict
from pathlib import Path


FRONTEND_ROOT = Path(__file__).resolve().parents[1]
BACKEND_ROOT = FRONTEND_ROOT.parents[1] / "ngoma_charts_backend" / "backend"
ARCHIVE = Path(r"C:\Users\HP\Downloads\June%202026%20Weeks%201-3%20Cleaned%20Files.zip")
MONTH_LABEL = "June 2026"
YEAR = 2026
MONTH = 6

sys.path.insert(0, str(BACKEND_ROOT))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "ngoma_backend.settings")

import django

django.setup()

import openpyxl
from django.apps import apps
from django.core.cache import cache
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify

from charts.artist_credits import format_artist_list, release_credit_payload
from charts.artist_metadata import artist_country
from charts.cms_utils import recalculate_certifications
from charts.master_dataset import (
    _artist_key,
    _canonical_artist_name,
    _normalize_title,
    _split_artist_credit,
)
from charts.models import (
    Artist,
    ChartType,
    MonthlyChart,
    MonthlyChartEntry,
    Platform,
    PlatformChartEntry,
    Release,
    ReleaseArtistCredit,
    WeeklyUpload,
)


PLATFORM_LIMITS = {
    "singles": {
        "Apple Music": 100,
        "Audiomack": 100,
        "Boomplay": 100,
        "Spotify": 50,
        "YouTube": 100,
        "Shazam": 100,
    },
    "albums": {
        "Apple Music": 200,
        "Audiomack": 200,
    },
}

TITLE_KEY_ALIASES = {
    "chaiyasaaumi": "chaiyasaakumi",
    "hallelujahzimbabwegospelsong2026": "hallelujah",
}


def clean_text(value):
    return " ".join(unicodedata.normalize("NFKC", str(value or "")).strip().split())


def simple_key(value):
    value = unicodedata.normalize("NFKD", clean_text(value))
    value = value.encode("ascii", "ignore").decode("ascii").casefold()
    return re.sub(r"[^a-z0-9]+", "", value)


def title_key(value):
    value = clean_text(value)
    value = re.sub(
        r"\s*[\(\[]\s*(?:feat|ft|featuring|with)\b.*?[\)\]]\s*$",
        "",
        value,
        flags=re.I,
    )
    value = re.sub(r"\s+(?:feat|ft|featuring|with)\.?\s+.*$", "", value, flags=re.I)
    value = re.sub(r"\s*-\s*(?:single|ep|live)\s*$", "", value, flags=re.I)
    key = simple_key(value)
    return TITLE_KEY_ALIASES.get(key, key)


def split_entry(raw, is_album):
    raw = clean_text(raw)
    if raw == "Kudade - Fancy Fingers Refix - Fancy Fingers":
        return "Kudade (Fancy Fingers Refix)", "Fancy Fingers"
    if raw.upper() == "BAHATI - CHERIE":
        return "Cherie", "Bahati"
    if " - " not in raw:
        return raw, ""
    index = raw.rfind(" - ") if is_album else raw.find(" - ")
    return raw[:index].strip(), raw[index + 3 :].strip()


def extract_title_features(title):
    features = []
    for match in re.finditer(
        r"[\(\[]\s*(?:feat|ft|featuring|with)\.?\s+([^\)\]]+)[\)\]]",
        title,
        flags=re.I,
    ):
        features.extend(_split_artist_credit(match.group(1)))
    trailing = re.search(r"\s+(?:feat|ft|featuring|with)\.?\s+(.+)$", title, flags=re.I)
    if trailing:
        features.extend(_split_artist_credit(trailing.group(1)))
    stripped = re.sub(
        r"\s*[\(\[]\s*(?:feat|ft|featuring|with)\b.*?[\)\]]\s*$",
        "",
        title,
        flags=re.I,
    )
    stripped = re.sub(
        r"\s+(?:feat|ft|featuring|with)\.?\s+.*$",
        "",
        stripped,
        flags=re.I,
    )
    return clean_text(stripped), unique_names(features)


def unique_names(values):
    result = []
    seen = set()
    for value in values:
        name = _canonical_artist_name(clean_text(value))
        key = _artist_key(name)
        if key and key not in seen and key != "unknown":
            result.append(name)
            seen.add(key)
    return result


def week_from_name(name):
    match = re.search(r"Week\s+(\d+)", name, re.I)
    return int(match.group(1)) if match else None


def chart_type_from_name(name):
    return "albums" if "Albums" in name else "singles"


def record_key(record):
    return title_key(record["title"]), _artist_key(record["primary"])


def record_members(record):
    return {
        _artist_key(name)
        for name in [record["primary"], *record["featured"]]
        if _artist_key(name)
    }


def load_catalog():
    by_type_title = {
        "singles": defaultdict(list),
        "albums": defaultdict(list),
    }
    entries = (
        MonthlyChartEntry.objects.filter(
            chart__is_published=True,
            chart__status="published",
            platform__isnull=True,
            rank__range=(1, 50),
        )
        .select_related("release", "release__artist")
        .prefetch_related("release__artist_credits__artist")
        .order_by("chart__year", "chart__month", "rank")
    )
    records_by_release = {}
    for entry in entries:
        release = entry.release
        payload = release_credit_payload(release, entry_featured=entry.featured_artists)
        record = records_by_release.get(release.id)
        if record is None:
            record = {
                "release_id": release.id,
                "title": release.title,
                "primary": payload["primary_artist_names"][0]
                if payload["primary_artist_names"]
                else release.artist.name,
                "featured": list(payload["featured_artist_names"]),
                "release_year": release.release_year or entry.release_year,
                "confidence": entry.confidence or "estimated (verify)",
                "existing": True,
            }
            records_by_release[release.id] = record
            by_type_title[release.chart_type][title_key(release.title)].append(record)
        else:
            record["featured"] = unique_names(
                [*record["featured"], *payload["featured_artist_names"]]
            )
            record["release_year"] = record["release_year"] or entry.release_year
            if entry.confidence == "web-verified":
                record["confidence"] = entry.confidence
    return by_type_title


def match_record(candidates, raw_names):
    raw_keys = {_artist_key(name) for name in raw_names if _artist_key(name)}
    raw_primary = _artist_key(raw_names[0]) if raw_names else ""
    scored = []
    for candidate in candidates:
        members = record_members(candidate)
        primary = _artist_key(candidate["primary"])
        overlap = len(raw_keys & members)
        if raw_keys and raw_keys == members:
            score = 1000
        elif raw_keys and raw_keys.issubset(members):
            score = 800 + overlap
        elif raw_primary and raw_primary == primary:
            score = 600 + overlap
        elif overlap:
            score = 100 + overlap
        elif raw_primary in ("", "unknown"):
            score = 10
        else:
            score = 0
        if score:
            scored.append((score, candidate))
    if not scored:
        return None
    scored.sort(key=lambda item: item[0], reverse=True)
    if len(scored) > 1 and scored[0][0] == scored[1][0]:
        return None
    return scored[0][1]


def load_workbook_rows():
    report_map = {}
    workbook_files = []
    with zipfile.ZipFile(ARCHIVE) as archive:
        for name in archive.namelist():
            if "Harmonization Report" in name:
                workbook = openpyxl.load_workbook(
                    io.BytesIO(archive.read(name)), read_only=True, data_only=True
                )
                sheet = workbook["Credit Harmonization"]
                for row in list(sheet.iter_rows(values_only=True))[1:]:
                    if row[0] and row[1] and row[3] and row[4]:
                        report_map[
                            (
                                clean_text(row[0]).casefold(),
                                clean_text(row[1]),
                                clean_text(row[3]),
                            )
                        ] = clean_text(row[4])
                workbook.close()
            else:
                workbook_files.append((name, archive.read(name)))

    rows = []
    for name, payload in workbook_files:
        chart_type = chart_type_from_name(name)
        week = week_from_name(name)
        workbook = openpyxl.load_workbook(
            io.BytesIO(payload), read_only=True, data_only=True
        )
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        headers = [clean_text(value) for value in values[0]]
        for platform_index, platform in enumerate(headers):
            limit = PLATFORM_LIMITS[chart_type].get(platform)
            if not limit:
                continue
            position = 0
            for source_row, row in enumerate(values[1:], 2):
                cell = row[platform_index] if platform_index < len(row) else None
                if cell in (None, ""):
                    continue
                position += 1
                if position > limit:
                    break
                raw = clean_text(cell)
                mapped = raw
                if week == 3:
                    mapped = report_map.get((chart_type, platform, raw), raw)
                rows.append(
                    {
                        "chart_type": chart_type,
                        "week": week,
                        "platform": platform,
                        "position": position,
                        "points": limit + 1 - position,
                        "raw": raw,
                        "mapped": mapped,
                        "source_file": Path(name).name,
                        "source_row": source_row,
                    }
                )
        workbook.close()
    return rows, report_map


def resolve_rows(raw_rows, catalog):
    june_catalog = {
        "singles": defaultdict(list),
        "albums": defaultdict(list),
    }
    resolution_counts = defaultdict(int)
    warnings = []

    for row in raw_rows:
        chart_type = row["chart_type"]
        is_album = chart_type == "albums"
        title, artist_credit = split_entry(row["mapped"], is_album)
        title, title_features = extract_title_features(title)
        artist_names = unique_names(_split_artist_credit(artist_credit))
        raw_names = unique_names([*artist_names, *title_features])
        if not raw_names:
            raw_names = ["Unknown Artist"]

        key = title_key(title)
        record = match_record(catalog[chart_type].get(key, []), raw_names)
        method = "existing"
        if record is None:
            record = match_record(june_catalog[chart_type].get(key, []), raw_names)
            method = "june"
        if record is None:
            record = {
                "release_id": None,
                "title": title,
                "primary": raw_names[0],
                "featured": unique_names(raw_names[1:]),
                "release_year": None,
                "confidence": "estimated (verify)",
                "existing": False,
            }
            june_catalog[chart_type][key].append(record)
            method = "new"
        elif (
            not record["existing"]
            and raw_names
            and _artist_key(raw_names[0]) == _artist_key(record["primary"])
        ):
            record["featured"] = unique_names([*record["featured"], *raw_names[1:]])

        resolution_counts[method] += 1
        row["record"] = record
        row["release_key"] = record_key(record)
        row["resolved"] = (
            f"{record['title']} - "
            f"{format_artist_list([record['primary'], *record['featured']])}"
        )
        if row["raw"] != row["resolved"] and method == "new":
            warnings.append(
                {
                    "kind": "new_variant",
                    "source": row["raw"],
                    "resolved": row["resolved"],
                    "platform": row["platform"],
                    "week": row["week"],
                }
            )
    return resolution_counts, warnings


def aggregate_rows(raw_rows):
    platform_groups = defaultdict(
        lambda: {
            "points": 0,
            "weeks": set(),
            "peak_position": 10_000,
            "record": None,
        }
    )
    for row in raw_rows:
        key = (row["chart_type"], row["platform"], row["release_key"])
        item = platform_groups[key]
        item["points"] += row["points"]
        item["weeks"].add(row["week"])
        item["peak_position"] = min(item["peak_position"], row["position"])
        item["record"] = row["record"]

    platform_rows = {"singles": [], "albums": []}
    for chart_type in ("singles", "albums"):
        for platform in PLATFORM_LIMITS[chart_type]:
            items = [
                (key, value)
                for key, value in platform_groups.items()
                if key[0] == chart_type and key[1] == platform
            ]
            items.sort(
                key=lambda item: (
                    -item[1]["points"],
                    item[1]["peak_position"],
                    item[1]["record"]["title"].casefold(),
                )
            )
            for rank, (_key, item) in enumerate(items, 1):
                record = item["record"]
                platform_rows[chart_type].append(
                    {
                        "month": MONTH_LABEL,
                        "platform": platform,
                        "rank": rank,
                        "title": record["title"],
                        "primary": record["primary"],
                        "featured": list(record["featured"]),
                        "points": item["points"],
                        "weeks": len(item["weeks"]),
                        "week_set": sorted(item["weeks"]),
                        "peak_position": item["peak_position"],
                        "record": record,
                        "release_key": record_key(record),
                    }
                )

    combined_rows = {"singles": [], "albums": []}
    for chart_type in ("singles", "albums"):
        combined = defaultdict(
            lambda: {
                "raw_points": 0,
                "platforms": set(),
                "weeks": set(),
                "record": None,
            }
        )
        for row in platform_rows[chart_type]:
            item = combined[row["release_key"]]
            item["raw_points"] += row["points"]
            item["platforms"].add(row["platform"])
            item["weeks"].update(row["week_set"])
            item["record"] = row["record"]
        ranked = list(combined.items())
        ranked.sort(
            key=lambda item: (
                -item[1]["raw_points"],
                -len(item[1]["platforms"]),
                item[1]["record"]["title"].casefold(),
            )
        )
        for rank, (release_key, item) in enumerate(ranked[:50], 1):
            record = item["record"]
            combined_rows[chart_type].append(
                {
                    "month": MONTH_LABEL,
                    "rank": rank,
                    "title": record["title"],
                    "primary": record["primary"],
                    "featured": list(record["featured"]),
                    "display_points": 51 - rank,
                    "raw_points": item["raw_points"],
                    "platform_count": len(item["platforms"]),
                    "platform_max": len(PLATFORM_LIMITS[chart_type]),
                    "weeks": len(item["weeks"]),
                    "record": record,
                    "release_key": release_key,
                }
            )
    return platform_rows, combined_rows


def serialize_summary(raw_rows, platform_rows, combined_rows, resolution_counts, warnings):
    return {
        "month": MONTH_LABEL,
        "source_archive": str(ARCHIVE),
        "raw_rows_processed": len(raw_rows),
        "resolution_counts": dict(resolution_counts),
        "warnings": warnings[:200],
        "weekly_counts": {
            chart_type: {
                str(week): sum(
                    1
                    for row in raw_rows
                    if row["chart_type"] == chart_type and row["week"] == week
                )
                for week in (1, 2, 3)
            }
            for chart_type in ("singles", "albums")
        },
        "combined": {
            chart_type: [
                {
                    key: (
                        format_artist_list(value)
                        if key == "featured"
                        else value
                    )
                    for key, value in row.items()
                    if key
                    in {
                        "rank",
                        "title",
                        "primary",
                        "featured",
                        "display_points",
                        "raw_points",
                        "platform_count",
                        "platform_max",
                        "weeks",
                    }
                }
                for row in combined_rows[chart_type]
            ]
            for chart_type in ("singles", "albums")
        },
        "platform_counts": {
            chart_type: {
                platform: sum(
                    1
                    for row in platform_rows[chart_type]
                    if row["platform"] == platform
                )
                for platform in PLATFORM_LIMITS[chart_type]
            }
            for chart_type in ("singles", "albums")
        },
    }


def safe_slug(name):
    base = slugify(name)[:45] or "unknown"
    candidate = base
    suffix = 2
    while Artist.objects.filter(slug=candidate).exclude(name=name).exists():
        candidate = f"{base[:40]}-{suffix}"
        suffix += 1
    return candidate


def get_artist(name):
    artist = Artist.objects.filter(name__iexact=name).first()
    if artist:
        return artist
    metadata = artist_country(name)
    defaults = {"slug": safe_slug(name)}
    if metadata:
        defaults.update(country=metadata[0], country_code=metadata[1])
    return Artist.objects.create(name=name, **defaults)


def get_release(record, chart_type):
    if record["release_id"]:
        release = Release.objects.filter(
            pk=record["release_id"], chart_type=chart_type
        ).first()
        if release:
            return release
    primary = get_artist(record["primary"])
    canonical_title = _normalize_title(record["title"])
    release = Release.objects.filter(
        chart_type=chart_type,
        canonical_title=canonical_title,
        artist=primary,
    ).first()
    if release is None:
        release = Release.objects.create(
            title=record["title"],
            canonical_title=canonical_title,
            artist=primary,
            chart_type=chart_type,
            featured_artists=format_artist_list(record["featured"]),
            release_year=record["release_year"],
        )
    if not release.featured_artists and record["featured"]:
        release.featured_artists = format_artist_list(record["featured"])
        release.save(update_fields=["featured_artists", "updated_at"])

    if not release.artist_credits.filter(role="primary").exists():
        ReleaseArtistCredit.objects.create(
            release=release, artist=primary, role="primary", position=0
        )
    existing_featured = {
        _artist_key(item.artist.name)
        for item in release.artist_credits.filter(role="featured").select_related("artist")
    }
    next_position = (
        release.artist_credits.filter(role="featured")
        .order_by("-position")
        .values_list("position", flat=True)
        .first()
        or -1
    ) + 1
    for name in record["featured"]:
        if _artist_key(name) in existing_featured:
            continue
        featured_artist = get_artist(name)
        ReleaseArtistCredit.objects.create(
            release=release,
            artist=featured_artist,
            role="featured",
            position=next_position,
        )
        existing_featured.add(_artist_key(name))
        next_position += 1
    return release


def previous_entry(release, chart_type, platform):
    return (
        MonthlyChartEntry.objects.filter(
            chart__year=2026,
            chart__month=5,
            chart__chart_type=chart_type,
            release=release,
            platform=platform,
        )
        .order_by("rank")
        .first()
    )


def historic_peak(release, chart_type, platform):
    value = (
        MonthlyChartEntry.objects.filter(
            chart__chart_type=chart_type,
            chart__year__lte=YEAR,
            release=release,
            platform=platform,
        )
        .order_by("rank")
        .values_list("rank", flat=True)
        .first()
    )
    return value or 999


@transaction.atomic
def apply_to_database(raw_rows, platform_rows, combined_rows):
    MonthlyChart.objects.filter(year=YEAR, month=MONTH).delete()
    WeeklyUpload.objects.filter(year=YEAR, month=MONTH).delete()

    platforms = {
        platform.name: platform
        for platform in Platform.objects.filter(
            name__in={
                name
                for limits in PLATFORM_LIMITS.values()
                for name in limits
            }
        )
    }
    release_cache = {}

    def release_for(record, chart_type):
        key = (chart_type, record_key(record))
        if key not in release_cache:
            release_cache[key] = get_release(record, chart_type)
        return release_cache[key]

    uploads = {}
    grouped_upload_rows = defaultdict(list)
    for row in raw_rows:
        grouped_upload_rows[(row["chart_type"], row["week"])].append(row)
    for (chart_type, week), rows in sorted(grouped_upload_rows.items()):
        source_names = sorted({row["source_file"] for row in rows})
        upload = WeeklyUpload.objects.create(
            chart_type=chart_type,
            year=YEAR,
            month=MONTH,
            week=week,
            file=source_names[0],
            processed=True,
            processing_notes=(
                f"Imported from {ARCHIVE.name}; canonicalized against published catalog."
            ),
            entries_processed=len(rows),
            duplicates_dropped=0,
        )
        uploads[(chart_type, week)] = upload
        entries = []
        for row in rows:
            title, raw_artist = split_entry(
                row["raw"], chart_type == ChartType.ALBUMS
            )
            entries.append(
                PlatformChartEntry(
                    upload=upload,
                    platform=platforms[row["platform"]],
                    release=release_for(row["record"], chart_type),
                    position=row["position"],
                    points=row["points"],
                    raw_title=title,
                    raw_artist=raw_artist,
                )
            )
        PlatformChartEntry.objects.bulk_create(entries, batch_size=500)

    created = {}
    for chart_type in ("singles", "albums"):
        chart = MonthlyChart.objects.create(
            year=YEAR,
            month=MONTH,
            chart_type=chart_type,
            label=MONTH_LABEL,
            status="published",
            is_published=True,
            locked=True,
            published_at=timezone.now(),
        )
        entries = []
        for row in platform_rows[chart_type]:
            release = release_for(row["record"], chart_type)
            platform = platforms[row["platform"]]
            previous = previous_entry(release, chart_type, platform)
            entries.append(
                MonthlyChartEntry(
                    chart=chart,
                    platform=platform,
                    release=release,
                    rank=row["rank"],
                    total_points=row["points"],
                    raw_total_points=None,
                    weeks_on_chart=row["weeks"],
                    platform_count=1,
                    platform_max=1,
                    featured_artists=format_artist_list(row["featured"]),
                    release_year=release.release_year,
                    confidence=row["record"]["confidence"],
                    peak_rank=min(
                        historic_peak(release, chart_type, platform), row["rank"]
                    ),
                    prev_rank=previous.rank if previous and previous.rank <= 50 else None,
                )
            )
        for row in combined_rows[chart_type]:
            release = release_for(row["record"], chart_type)
            previous = previous_entry(release, chart_type, None)
            entries.append(
                MonthlyChartEntry(
                    chart=chart,
                    platform=None,
                    release=release,
                    rank=row["rank"],
                    total_points=row["display_points"],
                    raw_total_points=row["raw_points"],
                    weeks_on_chart=row["weeks"],
                    platform_count=row["platform_count"],
                    platform_max=row["platform_max"],
                    featured_artists=format_artist_list(row["featured"]),
                    release_year=release.release_year,
                    confidence=row["record"]["confidence"],
                    peak_rank=min(
                        historic_peak(release, chart_type, None), row["rank"]
                    ),
                    prev_rank=previous.rank if previous else None,
                )
            )
        MonthlyChartEntry.objects.bulk_create(entries, batch_size=500)
        created[chart_type] = {
            "combined": len(combined_rows[chart_type]),
            "platform": len(platform_rows[chart_type]),
        }

    certification_updates = sum(
        recalculate_certifications(chart_type=chart_type)
        for chart_type in ("singles", "albums")
    )
    cache.clear()
    return {
        "charts": created,
        "weekly_uploads": len(uploads),
        "weekly_entries": len(raw_rows),
        "certification_updates": certification_updates,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--compact", action="store_true")
    parser.add_argument(
        "--output",
        type=Path,
        default=FRONTEND_ROOT / ".tmp" / "june_2026_processed.json",
    )
    args = parser.parse_args()

    catalog = load_catalog()
    raw_rows, report_map = load_workbook_rows()
    resolution_counts, warnings = resolve_rows(raw_rows, catalog)
    platform_rows, combined_rows = aggregate_rows(raw_rows)
    summary = serialize_summary(
        raw_rows, platform_rows, combined_rows, resolution_counts, warnings
    )
    summary["harmonization_mappings"] = len(report_map)
    if args.apply:
        summary["database"] = apply_to_database(
            raw_rows, platform_rows, combined_rows
        )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if args.compact:
        print(
            json.dumps(
                {
                    "month": summary["month"],
                    "raw_rows_processed": summary["raw_rows_processed"],
                    "resolution_counts": summary["resolution_counts"],
                    "weekly_counts": summary["weekly_counts"],
                    "platform_counts": summary["platform_counts"],
                    "top_10": {
                        chart_type: summary["combined"][chart_type][:10]
                        for chart_type in ("singles", "albums")
                    },
                    "database": summary.get("database"),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
    else:
        print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
