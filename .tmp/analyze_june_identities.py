import io
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict

import openpyxl


ARCHIVE = r"C:\Users\HP\Downloads\June%202026%20Weeks%201-3%20Cleaned%20Files.zip"


def split_entry(raw, is_album):
    raw = re.sub(r"\s+", " ", str(raw or "")).strip()
    if " - " not in raw:
        return raw, ""
    index = raw.rfind(" - ") if is_album else raw.find(" - ")
    return raw[:index].strip(), raw[index + 3 :].strip()


def simplify(value):
    value = unicodedata.normalize("NFKD", str(value or ""))
    value = value.encode("ascii", "ignore").decode("ascii").lower()
    value = re.sub(
        r"\s*[\(\[]\s*(?:feat|ft|featuring|with)\b.*?[\)\]]\s*$",
        "",
        value,
    )
    value = re.sub(r"\s+(?:feat|ft|featuring|with)\.?\s+.*$", "", value)
    value = re.sub(r"\s*-\s*(?:single|ep|live)\s*$", "", value)
    return re.sub(r"[^a-z0-9]+", "", value)


def week_from_name(name):
    match = re.search(r"Week\s+(\d+)", name, re.I)
    return int(match.group(1)) if match else None


entries = {"singles": [], "albums": []}
report_map = {}

with zipfile.ZipFile(ARCHIVE) as archive:
    for name in archive.namelist():
        if "Harmonization Report" in name:
            workbook = openpyxl.load_workbook(
                io.BytesIO(archive.read(name)), read_only=True, data_only=True
            )
            sheet = workbook["Credit Harmonization"]
            for row in list(sheet.iter_rows(values_only=True))[1:]:
                if row[3] and row[4]:
                    report_map[str(row[3]).strip()] = str(row[4]).strip()
            workbook.close()
            continue
        chart_type = "albums" if "Albums" in name else "singles"
        workbook = openpyxl.load_workbook(
            io.BytesIO(archive.read(name)), read_only=True, data_only=True
        )
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value).strip() if value else "" for value in rows[0]]
        for platform_index, platform in enumerate(headers):
            for position, row in enumerate(rows[1:], 1):
                if platform_index >= len(row) or row[platform_index] in (None, ""):
                    continue
                raw = str(row[platform_index]).strip()
                title, artist = split_entry(raw, chart_type == "albums")
                entries[chart_type].append(
                    {
                        "week": week_from_name(name),
                        "platform": platform,
                        "position": position,
                        "raw": raw,
                        "title": title,
                        "artist": artist,
                        "title_key": simplify(title),
                    }
                )
        workbook.close()

print(f"harmonization mappings={len(report_map)}")
for chart_type, rows in entries.items():
    early = [row for row in rows if row["week"] in (1, 2)]
    week3 = [row for row in rows if row["week"] == 3]
    early_by_title = defaultdict(list)
    for row in early:
        early_by_title[row["title_key"]].append(row)

    unmatched = []
    ambiguous = []
    resolved = []
    for row in week3:
        mapped = report_map.get(row["raw"])
        candidates = early_by_title.get(row["title_key"], [])
        canonical = Counter(item["raw"] for item in candidates)
        if mapped:
            resolved.append(("report", row, mapped))
        elif len(canonical) == 1:
            resolved.append(("unique", row, canonical.most_common(1)[0][0]))
        elif canonical:
            top = canonical.most_common()
            if len(top) == 1 or top[0][1] > top[1][1]:
                resolved.append(("frequency", row, top[0][0]))
            else:
                ambiguous.append((row, top[:8]))
        else:
            unmatched.append(row)

    print(f"\n{chart_type.upper()}")
    print(
        f"  entries={len(rows)} early={len(early)} week3={len(week3)} "
        f"resolved={len(resolved)} ambiguous={len(ambiguous)} unmatched={len(unmatched)}"
    )
    print(
        "  resolution="
        + str(Counter(method for method, _, _ in resolved))
    )
    print("  sample changed resolutions:")
    changes = [
        (method, row["raw"], target)
        for method, row, target in resolved
        if row["raw"] != target
    ]
    for item in changes[:30]:
        print(f"    {item}")
    print("  sample ambiguous:")
    for row, candidates in ambiguous[:20]:
        print(f"    {row['raw']!r} -> {candidates}")
    print("  sample unmatched:")
    for row in unmatched[:30]:
        print(f"    {row['platform']} #{row['position']}: {row['raw']}")
