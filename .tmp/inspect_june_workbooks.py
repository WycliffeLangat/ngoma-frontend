import io
import zipfile
from collections import Counter

import openpyxl


ARCHIVE = r"C:\Users\HP\Downloads\June%202026%20Weeks%201-3%20Cleaned%20Files.zip"


with zipfile.ZipFile(ARCHIVE) as archive:
    for name in archive.namelist():
        workbook = openpyxl.load_workbook(
            io.BytesIO(archive.read(name)),
            read_only=True,
            data_only=True,
        )
        print(f"\nFILE {name}")
        print(f"  sheets={workbook.sheetnames}")
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            width = max((len(row) for row in rows), default=0)
            headers = [
                str(value).strip() if value is not None else ""
                for value in (rows[0] if rows else ())
            ]
            print(f"  sheet={sheet.title!r} rows={len(rows)} cols={width}")
            print(f"  headers={headers}")
            for index, header in enumerate(headers):
                values = [
                    str(row[index]).strip()
                    for row in rows[1:]
                    if index < len(row) and row[index] not in (None, "")
                ]
                duplicates = sum(count - 1 for count in Counter(values).values() if count > 1)
                print(
                    f"    {header or f'column_{index + 1}'}: "
                    f"nonempty={len(values)} duplicates={duplicates}"
                )
                if values:
                    print(f"      first={values[:3]}")
        workbook.close()
